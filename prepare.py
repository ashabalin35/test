import pandas as pd
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils.cell import get_column_letter
from openpyxl import Workbook


def write_log_ove(channel, partner, col, data1, data2):
    xls = False
    months = {1: "января", 2: "февраля", 3: "марта", 4: "апреля", 5: "мая", 6: "июня", 7: "июля",
              8: "августа", 9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"}

    partner = partner.upper()
    main_link = 'c:\\AzRun\\'
    year = data1[0:4]
    month = int(data1[5:7])
    if month <= 9:
        month = '0' + str(month)
    else:
        month = str(month)
    day = int(data1[8:])
    year2 = data2[0:4]
    month2 = int(data2[5:7])
    if month2 <= 9:
        month2 = '0' + str(month2)
    else:
        month2 = str(month2)
    day2 = int(data2[8:])
    file = False
    if (year == year2) and (month == month2) and (day <= day2):
        n = day2 - day
        main_link += year + '_' + month + '\\'
        filelink = 'C:\\FlaskApp\\App\\file\\'
        file = f'OVE_{partner}_{data1}_{data2}.xlsx'
        report = filelink + file
        try:
            data, time, name, ove, dur = [], [], [], [], []
            df_new = False
            for i in range(n + 1):
                if (day + i) <= 9:
                    filename = main_link + channel + '_' + year + month + '0' + str(day + i) + '.txt'
                else:
                    filename = main_link + channel + '_' + year + month + str(day + i) + '.txt'
                df = pd.read_csv(filename, sep="\t",
                                 names=['data', 'time', 'type', 'name', 'dur', 'ch', '7', '8', '9', '10', '11',
                                        '12',
                                        '13',
                                        'ID',
                                        '15', 'ad_n', 'ad_dur', '18', '19', '20', '21', '22', '23', '24', '25',
                                        '26'])
                fut = ['data', 'time', 'type', 'name', 'dur', 'ch', 'ID', 'ad_n', 'ad_dur']
                df = df[fut]
                df_caps = df.copy()
                df_caps['name'] = df_caps['name'].apply(str.upper)
                df_caps['ID'] = df_caps['ID'].apply(str.upper)
                indexes = df_caps.index[(df_caps[col].str.contains(partner))].tolist()
                if len(indexes) != 0:
                    for index in indexes:
                        if df.iloc[index]['type'] != 'START':
                            for j in range(1, 10, 1):
                                if df.iloc[(index - j)]['type'] == 'START':
                                    data.append(df.iloc[index]['data'])
                                    time.append(df.iloc[index]['time'])
                                    name.append(df.iloc[index - j]['name'])
                                    ove.append(df.iloc[index]['name'])
                                    dur.append(df.iloc[index]['dur'][6:8])
                                    break
            if len(data) != 0:
                data_dic = {'Дата': data, 'Время': time, 'Название': name, 'Оверлей': ove, 'Хрон.(сек)': dur}
                df_new = pd.DataFrame(data_dic)
            df = df_new
            if df is not False:
                wb = Workbook()
                # ws = wb.active
                with pd.ExcelWriter(report, engine='openpyxl') as writer:
                    writer.book = wb
                    width = {}
                    for column in df:
                        column_width = max(df[column].astype(str).map(len).max(), len(column))
                        col_idx = df.columns.get_loc(column)
                        width[col_idx] = column_width
                    font = Font(
                        name='Calibri',
                        size=10,
                        bold=False,
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False,
                        color='00000000'
                    )
                    alignment = Alignment(
                        horizontal='general',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0
                    )
                    df.to_excel(writer, sheet_name='report', startrow=3, index=False, header=True)
                    ws = writer.sheets['report']
                    ws.merge_cells('A2:E2')
                    if n == 0:
                        ws['A2'] = f'Эфирная справка по оверлею "{partner}" за {day} {months[int(month)]} {year} года'
                    else:
                        ws[
                            'A2'] = f'Эфирная справка по оверлею "{partner}" за {day} - {day2} {months[int(month)]} {year} года'
                    ws['A2'].font = Font(name='Calibri', size=12, bold=True)
                    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
                    cell_range = 'A4:E' + str(ws.max_row)
                    set_style(ws, cell_range, font, alignment)
                    for i in range(len(width)):
                        letter = get_column_letter(i + 1)
                        ws.column_dimensions[letter].width = width[i]
                message = f'Данные для "{partner}" на канале {channel} за указанный период успешно сохранены'
            else:
                message = f'Данные для "{partner}" на канале {channel} за указанный период не найдены!'
                file = False

        except Exception as e:
            print(e)
            message = f'Возникла ошибка: {e}\nОбратитесь к Администратору'
            file = False

    else:
        message = 'Проверьте корректность выбранного диапазона дат'
    return message, file


def write_log_main(channel, partner, col, data1, data2):
    months = {1: "января", 2: "февраля", 3: "марта", 4: "апреля", 5: "мая", 6: "июня", 7: "июля",
              8: "августа", 9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"}

    partner = partner.upper()
    main_link = 'c:\\AzRun\\'
    year = data1[0:4]
    month = int(data1[5:7])
    if month <= 9:
        month = '0' + str(month)
    else:
        month = str(month)
    day = int(data1[8:])
    year2 = data2[0:4]
    month2 = int(data2[5:7])
    if month2 <= 9:
        month2 = '0' + str(month2)
    else:
        month2 = str(month2)
    day2 = int(data2[8:])
    file = False
    if (year == year2) and (month == month2) and (day <= day2):
        n = day2 - day
        main_link += year + '_' + month + '\\'
        filelink = 'C:\\FlaskApp\\App\\file\\'
        file = f'MAIN_{partner}_{data1}_{data2}.xlsx'
        report = filelink + file
        try:
            data, time, name, dur = [], [], [], []
            df_new = False
            for i in range(n + 1):
                if (day + i) <= 9:
                    filename = main_link + channel + '_' + year + month + '0' + str(day + i) + '.txt'
                else:
                    filename = main_link + channel + '_' + year + month + str(day + i) + '.txt'
                df = pd.read_csv(filename, sep="\t",
                                 names=['data', 'time', 'type', 'name', 'dur', 'ch', '7', '8', '9', '10', '11',
                                        '12',
                                        '13',
                                        'ID',
                                        '15', 'ad_n', 'ad_dur', '18', '19', '20', '21', '22', '23', '24', '25',
                                        '26'])
                fut = ['data', 'time', 'type', 'name', 'dur', 'ch', 'ID' ,'ad_n', 'ad_dur']
                df = df[fut]
                df_caps = df.copy()
                df_caps['name'] = df_caps['name'].apply(str.upper)
                df_caps['ID'] = df_caps['ID'].apply(str.upper)
                indexes = df_caps.index[(df_caps[col].str.contains(partner))].tolist()
                if len(indexes) != 0:
                    for index in indexes:
                        if df.iloc[index]['type'] == 'START':
                            data.append(df.iloc[index]['data'])
                            time.append(df.iloc[index]['time'])
                            name.append(df.iloc[index]['name'])
                            dur.append(df.iloc[index]['dur'][6:8])
            if len(data) != 0:
                data_dic = {'Дата': data, 'Время': time, 'Название': name, 'Хрон.(сек)': dur}
                df_new = pd.DataFrame(data_dic)
            df = df_new
            if df is not False:
                wb = Workbook()
                # ws = wb.active
                with pd.ExcelWriter(report, engine='openpyxl') as writer:
                    writer.book = wb
                    width = {}
                    for column in df:
                        column_width = max(df[column].astype(str).map(len).max(), len(column))
                        col_idx = df.columns.get_loc(column)
                        width[col_idx] = column_width
                    font = Font(
                        name='Calibri',
                        size=10,
                        bold=False,
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False,
                        color='00000000'
                    )
                    alignment = Alignment(
                        horizontal='general',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0
                    )
                    df.to_excel(writer, sheet_name='report', startrow=3, index=False, header=True)
                    ws = writer.sheets['report']
                    ws.merge_cells('A2:E2')
                    if n == 0:
                        ws['A2'] = f'Эфирная справка по событию "{partner}" за {day} {months[int(month)]} {year} года'
                    else:
                        ws[
                            'A2'] = f'Эфирная справка по событию "{partner}" за {day} - {day2} {months[int(month)]} {year} года'
                    ws['A2'].font = Font(name='Calibri', size=12, bold=True)
                    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
                    cell_range = 'A4:E' + str(ws.max_row)
                    set_style(ws, cell_range, font, alignment)
                    for i in range(len(width)):
                        letter = get_column_letter(i + 1)
                        ws.column_dimensions[letter].width = width[i]
                message = f'Данные для "{partner}" на канале {channel} за указанный период успешно сохранены'
            else:
                message = f'Данные для "{partner}" на канале {channel} за указанный период не найдены!'
                file = False

        except Exception as e:
            print(e)
            message = f'Возникла ошибка: {e}\nОбратитесь к Администратору'
            file = False

    else:
        message = 'Проверьте корректность выбранного диапазона дат'
    return message, file


def set_style(ws, cell_range, font, alignment):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.font = font
            cell.alignment = alignment
